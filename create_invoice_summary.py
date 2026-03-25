#!/usr/bin/env python3
"""Generate consolidated Excel with all invoices, each date as a column, sorted by code."""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import OrderedDict

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Invoice Summary"

# === Invoice Data (sorted by date) ===
invoices = [
    {
        'date': '01/02/2026',
        'invoice_no': 'DFM8101729',
        'driver': '陈桥桥',
        'note': 'AMENDED',
        'items': [
            ('BCS02',  '豆福豆皮（散装1KG） DOLFORD BEANCURD SKIN',            'KGS', 3.60,  2.00,   7.20),
            ('CST01',  '芝士海鲜豆腐 CHEESE SEAFOOD TOFU',                     'PKT', 3.50,  6.00,  21.00),
            ('ET01P',  '蛋豆腐 EGG TOFU 130g',                                'PKT', 0.25,  6.00,   1.50),
            ('EZ01C',  'EZ快热面 E-ZEE INSTANT NOODLES',                       'CTN', 18.00, 1.00,  18.00),
            ('FD01P',  '炸豆枝片 EB FRIED BEAN CURD SHEET',                    'PKT', 9.30,  2.00,  18.60),
            ('LB02P',  'FG 龙虾球 LOBSTER BALL（FG）',                          'PKT', 3.50,  3.00,  10.50),
            ('NTU01',  '豆福手工卤水大豆腐 DOLFORD NORTH TOFU',                  'PCS', 1.20,  4.00,   4.80),
            ('SAT01',  '豆福五香豆干（散装）1KG DOLFORD SPICED',                  'KGS', 3.60,  1.00,   3.60),
            ('SDO01',  '烟熏鸭胸原味 Smoked Duck Original 5PKT/KG',             'KGS', 9.50,  1.27,  12.07),
            ('VTC02',  '豆福素鸡（散装） DOLFORD VEGETARIAN',                    'KGS', 3.60,  1.00,   3.60),
            ('WFB01',  '白鱼丸 (FG)_1KG*10PKT/CTN WHITE FISH',                'KGS', 5.00,  1.00,   5.00),
            ('ZSB0080','金包银 MUSHROOM FISHN SOY 300G*30PKT/CTN',             'PKT', 2.50,  2.00,   5.00),
            ('NTU01',  'Exchange 交换：豆福手工卤水大豆腐 DOLFORD',               'PCS', None,  4.00,   None),
            ('DB01P',  '豆卜 DARK BROWN BEAN CURD 10PCS X 200G',              'PKT', 1.40, 15.00,  21.00),
        ],
        'subtotal': 131.87,
        'gst': 11.87,
        'total': 143.74,
    },
    {
        'date': '07/02/2026',
        'invoice_no': 'DFM8102534',
        'driver': '顾洪亮',
        'note': '',
        'items': [
            ('BCS02',  '豆福豆皮（散装1KG） DOLFORD BEANCURD SKIN',            'KGS', 3.60,  2.00,   7.20),
            ('CST01',  '芝士海鲜豆腐 CHEESE SEAFOOD TOFU',                     'PKT', 3.50,  5.00,  17.50),
            ('ET01P',  '蛋豆腐 EGG TOFU 130g',                                'PKT', 0.25,  6.00,   1.50),
            ('LB02P',  'FG 龙虾球 LOBSTER BALL（FG）',                          'PKT', 3.50,  3.00,  10.50),
            ('NTU01',  '豆福手工卤水大豆腐 DOLFORD NORTH TOFU',                  'PCS', 1.20,  8.00,   9.60),
            ('SDO01',  '烟熏鸭胸原味 Smoked Duck Original 5PKT/KG',             'KGS', 9.50,  1.14,  10.83),
            ('ZSB0080','金包银 MUSHROOM FISHN SOY 300G*30PKT/CTN',             'PKT', 2.65,  2.00,   5.30),
            ('DB01P',  '豆卜 DARK BROWN BEAN CURD 10PCS X 200G',              'PKT', 1.40, 10.00,  14.00),
            ('SAT01',  '豆福五香豆干（散装）1KG DOLFORD SPICED',                  'KGS', 3.60,  1.00,   3.60),
            ('VYC03',  '豆福素鸡（散装） DOLFORD VEGETARIAN',                    'KGS', 3.60,  1.00,   3.60),
            ('ZSB0082','蟹棒 MUSHROOM IMI.CRAB STICK',                         'PKT', 1.70, 15.00,  25.50),
        ],
        'subtotal': 109.13,
        'gst': 9.82,
        'total': 118.95,
    },
    {
        'date': '08/02/2026',
        'invoice_no': 'DFM8102644',
        'driver': '顾洪亮',
        'note': '',
        'items': [
            ('BCS02',  '豆福豆皮（散装1KG） DOLFORD BEANCURD SKIN',            'KGS', 3.60,  2.00,   7.20),
            ('CST01',  '芝士海鲜豆腐 CHEESE SEAFOOD TOFU',                     'PKT', 3.50,  5.00,  17.50),
            ('ET01P',  '蛋豆腐 EGG TOFU 130g',                                'PKT', 0.25, 10.00,   2.50),
            ('FD01P',  '炸豆枝片 EB FRIED BEAN CURD SHEET',                    'PKT', 9.30,  1.00,   9.30),
            ('LB02P',  'FG 龙虾球 LOBSTER BALL（FG）',                          'PKT', 3.50,  2.00,   7.00),
            ('NTU01',  '豆福手工卤水大豆腐 DOLFORD NORTH TOFU',                  'PCS', 1.20,  4.00,   4.80),
            ('SDO01',  '烟熏鸭胸原味 Smoked Duck Original 5PKT/KG',             'KGS', 9.50,  1.26,  11.97),
            ('WDM010', '日式乌冬面 INSTANT JAPANESE FRESH UDON',               'CTN', 21.00, 1.00,  21.00),
            ('WFB01',  '白鱼丸 (FG)_1KG*10PKT/CTN WHITE FISH',                'KGS', 5.00,  1.00,   5.00),
            ('ZSB0080','金包银 MUSHROOM FISHN SOY 300G*30PKT/CTN',             'PKT', 2.65,  1.00,   2.65),
        ],
        'subtotal': 88.92,
        'gst': 8.00,
        'total': 96.92,
    },
]

# Collect all unique products (code -> description, unit, price)
all_products = OrderedDict()
for inv in invoices:
    for code, desc, unit, price, qty, amt in inv['items']:
        if code not in all_products:
            all_products[code] = (desc, unit, price)

# Sort by code
sorted_codes = sorted(all_products.keys())

# Styles
title_font = Font(name='Arial', size=14, bold=True)
header_font = Font(name='Arial', size=10, bold=True)
normal_font = Font(name='Arial', size=10)
small_font = Font(name='Arial', size=9)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(name='Arial', size=10, bold=True, color='FFFFFF')
date_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
subtotal_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

# Column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 10

# Date columns: each date has 2 sub-columns (QTY, AMOUNT)
num_dates = len(invoices)
for i in range(num_dates):
    col_qty = chr(ord('F') + i * 2)
    col_amt = chr(ord('G') + i * 2)
    ws.column_dimensions[col_qty].width = 10
    ws.column_dimensions[col_amt].width = 12

# Last column for total
total_col_idx = 6 + num_dates * 2
total_col = chr(ord('A') + total_col_idx - 1)  # 'L'
ws.column_dimensions[total_col].width = 14

# === Title ===
row = 1
last_col = chr(ord('A') + total_col_idx - 1)
ws.merge_cells(f'A1:{last_col}1')
ws['A1'] = '豆福食品制造有限公司 DOLFORD FOOD MANUFACTURING PTE LTD - Invoice Summary'
ws['A1'].font = title_font
ws['A1'].alignment = center

ws.merge_cells(f'A2:{last_col}2')
ws['A2'] = 'Customer: ZHANG LIANG MALA TANG 张亮麻辣烫 (远东广场)  |  Customer ID: ZLML07  |  Term: 30 DAYS'
ws['A2'].font = Font(name='Arial', size=10)
ws['A2'].alignment = center

# === Header Row 1: Date groups ===
row = 4

# Fixed columns header
fixed_headers_r1 = ['No.', 'Code', 'ITEM NAME / DESCRIPTION 货名/摘要', 'UNIT 单位', 'U_PRICE 单价']
for col_idx, h in enumerate(fixed_headers_r1, 1):
    cell = ws.cell(row=row, column=col_idx, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

# Date group headers
for i, inv in enumerate(invoices):
    col_start = 6 + i * 2
    col_end = col_start + 1
    col_start_letter = chr(ord('A') + col_start - 1)
    col_end_letter = chr(ord('A') + col_end - 1)
    ws.merge_cells(f'{col_start_letter}{row}:{col_end_letter}{row}')
    cell = ws.cell(row=row, column=col_start)
    label = f"{inv['date']}\n{inv['invoice_no']}"
    if inv['note']:
        label += f"\n({inv['note']})"
    cell.value = label
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Border for merged cell right side
    ws.cell(row=row, column=col_end).border = thin_border

# Total column header
cell = ws.cell(row=row, column=total_col_idx, value='TOTAL\n合计')
cell.font = header_font_white
cell.fill = header_fill
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.row_dimensions[row].height = 45

# === Header Row 2: Sub-headers ===
row = 5
sub_headers_fixed = ['', '', '', '', '']
for col_idx, h in enumerate(sub_headers_fixed, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.border = thin_border
    cell.fill = date_fill

for i in range(num_dates):
    col_qty = 6 + i * 2
    col_amt = col_qty + 1
    cell_q = ws.cell(row=row, column=col_qty, value='QTY 数量')
    cell_q.font = header_font; cell_q.fill = date_fill; cell_q.border = thin_border; cell_q.alignment = center
    cell_a = ws.cell(row=row, column=col_amt, value='AMOUNT 金额')
    cell_a.font = header_font; cell_a.fill = date_fill; cell_a.border = thin_border; cell_a.alignment = center

cell_t = ws.cell(row=row, column=total_col_idx, value='AMOUNT 金额')
cell_t.font = header_font; cell_t.fill = date_fill; cell_t.border = thin_border; cell_t.alignment = center

# === Data Rows ===
row = 6
for idx, code in enumerate(sorted_codes):
    desc, unit, price = all_products[code]
    r = row + idx
    ws.row_dimensions[r].height = 20

    ws.cell(row=r, column=1, value=idx + 1).font = normal_font
    ws.cell(row=r, column=1).alignment = center
    ws.cell(row=r, column=1).border = thin_border

    ws.cell(row=r, column=2, value=code).font = normal_font
    ws.cell(row=r, column=2).alignment = left_align
    ws.cell(row=r, column=2).border = thin_border

    ws.cell(row=r, column=3, value=desc).font = normal_font
    ws.cell(row=r, column=3).alignment = left_align
    ws.cell(row=r, column=3).border = thin_border

    ws.cell(row=r, column=4, value=unit).font = normal_font
    ws.cell(row=r, column=4).alignment = center
    ws.cell(row=r, column=4).border = thin_border

    price_cell = ws.cell(row=r, column=5, value=price)
    price_cell.font = normal_font
    price_cell.alignment = right_align
    price_cell.border = thin_border
    price_cell.number_format = '#,##0.00'

    row_total = 0
    for i, inv in enumerate(invoices):
        col_qty = 6 + i * 2
        col_amt = col_qty + 1

        # Find this code in this invoice
        found_qty = None
        found_amt = None
        for ic, idesc, iunit, iprice, iqty, iamt in inv['items']:
            if ic == code:
                found_qty = iqty
                found_amt = iamt
                break

        cell_q = ws.cell(row=r, column=col_qty)
        cell_q.border = thin_border
        cell_q.alignment = center
        cell_q.font = normal_font

        cell_a = ws.cell(row=r, column=col_amt)
        cell_a.border = thin_border
        cell_a.alignment = right_align
        cell_a.font = normal_font
        cell_a.number_format = '#,##0.00'

        if found_qty is not None:
            cell_q.value = found_qty
            cell_q.number_format = '0.00'
        if found_amt is not None:
            cell_a.value = found_amt
            row_total += found_amt

    # Total column
    total_cell = ws.cell(row=r, column=total_col_idx, value=row_total if row_total > 0 else None)
    total_cell.font = Font(name='Arial', size=10, bold=True)
    total_cell.alignment = right_align
    total_cell.border = thin_border
    total_cell.number_format = '#,##0.00'

# === Summary Rows ===
summary_start = row + len(sorted_codes) + 1

# SUB-TOTAL
r = summary_start
ws.merge_cells(f'A{r}:E{r}')
ws.cell(row=r, column=1, value='金额 SUB-TOTAL').font = header_font
ws.cell(row=r, column=1).alignment = right_align
ws.cell(row=r, column=1).fill = subtotal_fill
ws.cell(row=r, column=1).border = thin_border
for c in range(2, 6):
    ws.cell(row=r, column=c).border = thin_border
    ws.cell(row=r, column=c).fill = subtotal_fill

grand_subtotal = 0
for i, inv in enumerate(invoices):
    col_qty = 6 + i * 2
    col_amt = col_qty + 1
    ws.cell(row=r, column=col_qty).border = thin_border
    ws.cell(row=r, column=col_qty).fill = subtotal_fill
    cell = ws.cell(row=r, column=col_amt, value=inv['subtotal'])
    cell.font = header_font; cell.alignment = right_align
    cell.border = thin_border; cell.number_format = '$#,##0.00'
    cell.fill = subtotal_fill
    grand_subtotal += inv['subtotal']

cell = ws.cell(row=r, column=total_col_idx, value=grand_subtotal)
cell.font = Font(name='Arial', size=10, bold=True)
cell.alignment = right_align; cell.border = thin_border
cell.number_format = '$#,##0.00'; cell.fill = subtotal_fill

# GST
r = summary_start + 1
ws.merge_cells(f'A{r}:E{r}')
ws.cell(row=r, column=1, value='另加消费税9% Add GST 9%').font = header_font
ws.cell(row=r, column=1).alignment = right_align
ws.cell(row=r, column=1).fill = subtotal_fill
ws.cell(row=r, column=1).border = thin_border
for c in range(2, 6):
    ws.cell(row=r, column=c).border = thin_border
    ws.cell(row=r, column=c).fill = subtotal_fill

grand_gst = 0
for i, inv in enumerate(invoices):
    col_qty = 6 + i * 2
    col_amt = col_qty + 1
    ws.cell(row=r, column=col_qty).border = thin_border
    ws.cell(row=r, column=col_qty).fill = subtotal_fill
    cell = ws.cell(row=r, column=col_amt, value=inv['gst'])
    cell.font = header_font; cell.alignment = right_align
    cell.border = thin_border; cell.number_format = '$#,##0.00'
    cell.fill = subtotal_fill
    grand_gst += inv['gst']

cell = ws.cell(row=r, column=total_col_idx, value=round(grand_gst, 2))
cell.font = Font(name='Arial', size=10, bold=True)
cell.alignment = right_align; cell.border = thin_border
cell.number_format = '$#,##0.00'; cell.fill = subtotal_fill

# TOTAL
r = summary_start + 2
ws.merge_cells(f'A{r}:E{r}')
ws.cell(row=r, column=1, value='总金额 TOTAL').font = Font(name='Arial', size=11, bold=True)
ws.cell(row=r, column=1).alignment = right_align
ws.cell(row=r, column=1).fill = total_fill
ws.cell(row=r, column=1).border = thin_border
for c in range(2, 6):
    ws.cell(row=r, column=c).border = thin_border
    ws.cell(row=r, column=c).fill = total_fill

grand_total = 0
for i, inv in enumerate(invoices):
    col_qty = 6 + i * 2
    col_amt = col_qty + 1
    ws.cell(row=r, column=col_qty).border = thin_border
    ws.cell(row=r, column=col_qty).fill = total_fill
    cell = ws.cell(row=r, column=col_amt, value=inv['total'])
    cell.font = Font(name='Arial', size=11, bold=True); cell.alignment = right_align
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='double'))
    cell.number_format = '$#,##0.00'; cell.fill = total_fill
    grand_total += inv['total']

cell = ws.cell(row=r, column=total_col_idx, value=round(grand_total, 2))
cell.font = Font(name='Arial', size=12, bold=True)
cell.alignment = right_align
cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='double'))
cell.number_format = '$#,##0.00'; cell.fill = total_fill

# Print setup
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1

output_path = '/home/user/JoeyZzz/invoice_summary.xlsx'
wb.save(output_path)
print(f'Summary saved to {output_path}')
