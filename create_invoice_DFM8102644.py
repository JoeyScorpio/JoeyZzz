#!/usr/bin/env python3
"""Generate Excel for invoice DFM8102644."""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tax Invoice"

col_widths = {'A': 5, 'B': 12, 'C': 45, 'D': 8, 'E': 8, 'F': 10, 'G': 12}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

title_font = Font(name='Arial', size=16, bold=True)
header_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
small_font = Font(name='Arial', size=9)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')
wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# === Company Header ===
ws.merge_cells('A1:C3')
ws['A1'] = '豆福食品制造有限公司\nDOLFORD FOOD MANUFACTURING PTE LTD\n8A Admiralty Street #02-17/18 FoodXchange@Admiralty (S)757437'
ws['A1'].font = Font(name='Arial', size=11, bold=True)
ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws.merge_cells('D1:E1')
ws['D1'] = 'TAX INVOICE'
ws['D1'].font = title_font
ws['D1'].alignment = center

ws.merge_cells('F1:G1')
ws['F1'] = 'GST Reg No:201010861E'
ws['F1'].font = small_font
ws['F1'].alignment = right_align

ws.merge_cells('D2:E2')
ws['D2'] = 'INVOICE NO.:'
ws['D2'].font = header_font
ws['D2'].alignment = right_align

ws.merge_cells('F2:G2')
ws['F2'] = 'DFM8102644'
ws['F2'].font = header_font

ws.merge_cells('D3:E3')
ws['D3'] = 'DATE:'
ws['D3'].font = header_font
ws['D3'].alignment = right_align

ws.merge_cells('F3:G3')
ws['F3'] = '08/02/2026'
ws['F3'].font = header_font

# === Contact Info ===
ws.merge_cells('A4:G4')
ws['A4'] = 'TEL:(65)65706728  HP:(65)91784466  FAX:(65)65706729  Email:newdolford@yahoo.com'
ws['A4'].font = small_font
ws['A4'].alignment = center

ws.merge_cells('A5:G5')
ws['A5'] = 'Company UEN No.:201010861E  OCBC:509733820001'
ws['A5'].font = small_font
ws['A5'].alignment = center

# === Bill To / Deliver To ===
ws.merge_cells('A6:C6')
ws['A6'] = 'BILL TO:'
ws['A6'].font = header_font
ws['A6'].fill = header_fill
ws['A6'].border = thin_border

ws.merge_cells('E6:G6')
ws['E6'] = 'DELIVER TO:'
ws['E6'].font = header_font
ws['E6'].fill = header_fill
ws['E6'].border = thin_border

bill_to = [
    'ZHANG LIANG MALA TANG',
    '张亮麻辣烫 (远东广场) (DELIVERY AFTER 10:00AM)',
    'FAR EAST PLAZA 远东广场',
    '14 SCOTTS RD #01-07A SINGAPORE 228213',
    '电话: 65310331',
]
deliver_to = [
    'ZHANG LIANG MALA TANG',
    '张亮麻辣烫 (远东广场) (DELIVERY AFTER 10:00AM)',
    'FAR EAST PLAZA 远东广场',
    '14 SCOTTS RD #01-07A SINGAPORE 228213',
    '电话: 65310331',
]
for i in range(5):
    r = 7 + i
    ws.merge_cells(f'A{r}:C{r}')
    ws[f'A{r}'] = bill_to[i]
    ws[f'A{r}'].font = Font(name='Arial', size=10, bold=(i == 0))
    ws[f'A{r}'].border = thin_border
    ws.merge_cells(f'E{r}:G{r}')
    ws[f'E{r}'] = deliver_to[i]
    ws[f'E{r}'].font = Font(name='Arial', size=10, bold=(i == 0))
    ws[f'E{r}'].border = thin_border

# === Customer Info ===
ws.merge_cells('A13:B13')
ws['A13'] = 'Customer ID:'
ws['A13'].font = header_font
ws['C13'] = 'ZLML07'
ws['C13'].font = normal_font

ws.merge_cells('D13:E13')
ws['D13'] = 'Term:'
ws['D13'].font = header_font
ws['D13'].alignment = right_align
ws.merge_cells('F13:G13')
ws['F13'] = '30 DAYS'
ws['F13'].font = normal_font

ws.merge_cells('A14:B14')
ws['A14'] = 'Driver:'
ws['A14'].font = header_font
ws['C14'] = '陈桥桥'
ws['C14'].font = normal_font

ws.merge_cells('D14:E14')
ws['D14'] = 'PO No.:'
ws['D14'].font = header_font
ws['D14'].alignment = right_align

# === Table Header ===
row = 16
headers = ['No.', 'Code', 'ITEM NAME / DESCRIPTION\n货名/摘要', 'QTY\n数量', 'UNIT\n单位', 'U_PRICE\n单价', 'AMOUNT($)\n金额']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    cell.fill = header_fill
ws.row_dimensions[row].height = 30

# === Invoice Data ===
items = [
    (1,  'BCS02',   '豆福豆皮（散装1KG） DOLFORD BEANCURD SKIN',          2.00,  'KGS', 3.60,   7.20),
    (2,  'CST01',   '芝士海鲜豆腐 CHEESE SEAFOOD TOFU',                   5.00,  'PKT', 3.50,  17.50),
    (3,  'ET01P',   '蛋豆腐 EGG TOFU 130g',                              10.00,  'PKT', 0.25,   2.50),
    (4,  'FD01P',   '炸豆枝片 EB FRIED BEAN CURD SHEET',                  1.00,  'PKT', 9.30,   9.30),
    (5,  'LB02P',   'FG 龙虾球 LOBSTER BALL（FG）',                        2.00,  'PKT', 3.50,   7.00),
    (6,  'NTU01',   '豆福手工卤水大豆腐 DOLFORD NORTH TOFU',                4.00,  'PCS', 1.20,   4.80),
    (7,  'SDO01',   '烟熏鸭胸原味 Smoked Duck Original 5PKT/KG',           1.26,  'KGS', 9.50,  11.97),
    (8,  'WDM010',  '日式乌冬面 INSTANT JAPANESE FRESH UDON',              1.00,  'CTN', 21.00, 21.00),
    (9,  'WFB01',   '白鱼丸 (FG)_1KG*10PKT/CTN WHITE FISH',              1.00,  'KGS', 5.00,   5.00),
    (10, 'ZSB0080', '金包银 MUSHROOM FISHN SOY 300G*30PKT/CTN',           1.00,  'PKT', 2.65,   2.65),
]

for i, (no, code, desc, qty, unit, price, amount) in enumerate(items):
    r = 17 + i
    ws.row_dimensions[r].height = 20

    cell_no = ws.cell(row=r, column=1, value=no)
    cell_no.font = normal_font; cell_no.alignment = center; cell_no.border = thin_border

    cell_code = ws.cell(row=r, column=2, value=code)
    cell_code.font = normal_font; cell_code.alignment = left_align; cell_code.border = thin_border

    cell_desc = ws.cell(row=r, column=3, value=desc)
    cell_desc.font = normal_font; cell_desc.alignment = left_align; cell_desc.border = thin_border

    cell_qty = ws.cell(row=r, column=4, value=qty)
    cell_qty.font = normal_font; cell_qty.alignment = center; cell_qty.border = thin_border
    cell_qty.number_format = '0.00'

    cell_unit = ws.cell(row=r, column=5, value=unit)
    cell_unit.font = normal_font; cell_unit.alignment = center; cell_unit.border = thin_border

    cell_price = ws.cell(row=r, column=6, value=price)
    cell_price.font = normal_font; cell_price.alignment = right_align; cell_price.border = thin_border
    cell_price.number_format = '#,##0.00'

    cell_amt = ws.cell(row=r, column=7, value=amount)
    cell_amt.font = normal_font; cell_amt.alignment = right_align; cell_amt.border = thin_border
    cell_amt.number_format = '#,##0.00'

# Empty rows
last_data_row = 17 + len(items) - 1
for r in range(last_data_row + 1, 32):
    ws.row_dimensions[r].height = 20
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c)
        cell.font = normal_font
        cell.border = thin_border

# === Summary ===
summary_row = 33
sub_total = 88.92
gst = 8.00
total = 96.92

ws.merge_cells(f'A{summary_row}:E{summary_row}')
ws[f'F{summary_row}'] = '金额 SUB-TOTAL:'
ws[f'F{summary_row}'].font = header_font
ws[f'F{summary_row}'].alignment = right_align
ws[f'G{summary_row}'] = sub_total
ws[f'G{summary_row}'].font = header_font
ws[f'G{summary_row}'].alignment = right_align
ws[f'G{summary_row}'].number_format = '$#,##0.00'
ws[f'G{summary_row}'].border = thin_border

summary_row += 1
ws.merge_cells(f'A{summary_row}:E{summary_row}')
ws[f'F{summary_row}'] = '另加消费税9% Add GST 9%:'
ws[f'F{summary_row}'].font = header_font
ws[f'F{summary_row}'].alignment = right_align
ws[f'G{summary_row}'] = gst
ws[f'G{summary_row}'].font = header_font
ws[f'G{summary_row}'].alignment = right_align
ws[f'G{summary_row}'].number_format = '$#,##0.00'
ws[f'G{summary_row}'].border = thin_border

summary_row += 1
ws.merge_cells(f'A{summary_row}:E{summary_row}')
ws[f'F{summary_row}'] = '总金额 TOTAL:'
ws[f'F{summary_row}'].font = Font(name='Arial', size=12, bold=True)
ws[f'F{summary_row}'].alignment = right_align
ws[f'G{summary_row}'] = total
ws[f'G{summary_row}'].font = Font(name='Arial', size=12, bold=True)
ws[f'G{summary_row}'].alignment = right_align
ws[f'G{summary_row}'].number_format = '$#,##0.00'
ws[f'G{summary_row}'].border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='double')
)

# === Footer ===
note_row = 37
ws.merge_cells(f'A{note_row}:G{note_row}')
ws[f'A{note_row}'] = 'Note: Any Shortage in delivery, missing of goods, please notify the company between 9am to 6pm otherwise we would assume they are received according to order.'
ws[f'A{note_row}'].font = Font(name='Arial', size=8)
ws[f'A{note_row}'].alignment = wrap

note_row += 1
ws.merge_cells(f'A{note_row}:G{note_row}')
ws[f'A{note_row}'] = 'All cheque should be crossed and made payable to DOLFORD FOOD MANUFACTURING PTE LTD'
ws[f'A{note_row}'].font = Font(name='Arial', size=8)

note_row += 2
ws.merge_cells(f'A{note_row}:C{note_row}')
ws[f'A{note_row}'] = "Customer's Signature & Company Stamp"
ws[f'A{note_row}'].font = small_font
ws[f'A{note_row}'].border = Border(top=Side(style='thin'))

ws.print_area = 'A1:G40'
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1

output_path = '/home/user/JoeyZzz/invoice_DFM8102644.xlsx'
wb.save(output_path)
print(f'Invoice saved to {output_path}')
