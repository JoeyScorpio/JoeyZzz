#!/usr/bin/env python3
"""Generate a Tax Invoice Excel template based on Dolford Food Manufacturing invoice."""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tax Invoice"

# Column widths
col_widths = {'A': 5, 'B': 12, 'C': 45, 'D': 8, 'E': 8, 'F': 10, 'G': 12}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Styles
title_font = Font(name='Arial', size=16, bold=True)
header_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
small_font = Font(name='Arial', size=9)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')
right = Alignment(horizontal='right', vertical='center')
wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# === Company Header ===
ws.merge_cells('A1:C3')
ws['A1'] = '豆福食品制造有限公司\nDOLFORD FOOD MANUFACTURING PTE LTD\n8A Admiralty Street #02-17/18 FoodXchange@Admiralty (S)757437'
ws['A1'].font = Font(name='Arial', size=11, bold=True)
ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws.merge_cells('D1:E1')
ws['D1'] = 'TAX INVOICE'
ws['D1'].font = Font(name='Arial', size=16, bold=True)
ws['D1'].alignment = center

ws.merge_cells('F1:G1')
ws['F1'] = 'GST Reg No:'
ws['F1'].font = small_font
ws['F1'].alignment = right

ws.merge_cells('D2:E2')
ws['D2'] = 'INVOICE NO.:'
ws['D2'].font = header_font
ws['D2'].alignment = right

ws.merge_cells('F2:G2')
ws['F2'] = ''  # Invoice number placeholder
ws['F2'].font = header_font

ws.merge_cells('D3:E3')
ws['D3'] = 'DATE:'
ws['D3'].font = header_font
ws['D3'].alignment = right

ws.merge_cells('F3:G3')
ws['F3'] = ''  # Date placeholder
ws['F3'].font = header_font

# === Contact Info ===
row = 4
ws.merge_cells('A4:G4')
ws['A4'] = 'TEL:(65)65706728  HP:(65)91784466  FAX:(65)65706729  Email:newdolford@yahoo.com'
ws['A4'].font = small_font
ws['A4'].alignment = center

# === Bill To / Deliver To ===
row = 6
ws.merge_cells('A6:C6')
ws['A6'] = 'BILL TO:'
ws['A6'].font = header_font
ws['A6'].fill = header_fill

ws.merge_cells('E6:G6')
ws['E6'] = 'DELIVER TO:'
ws['E6'].font = header_font
ws['E6'].fill = header_fill

for r in range(7, 11):
    ws.merge_cells(f'A{r}:C{r}')
    ws[f'A{r}'].font = normal_font
    ws.merge_cells(f'E{r}:G{r}')
    ws[f'E{r}'].font = normal_font

# === Customer Info Row ===
row = 12
ws.merge_cells('A12:B12')
ws['A12'] = 'Customer ID:'
ws['A12'].font = header_font

ws['C12'] = ''  # Customer ID placeholder
ws['C12'].font = normal_font

ws.merge_cells('D12:E12')
ws['D12'] = 'Term:'
ws['D12'].font = header_font
ws['D12'].alignment = right

ws['F12'] = ''  # Term placeholder
ws['F12'].font = normal_font

ws.merge_cells('A13:B13')
ws['A13'] = 'Driver:'
ws['A13'].font = header_font

ws['C13'] = ''  # Driver placeholder
ws['C13'].font = normal_font

ws.merge_cells('D13:E13')
ws['D13'] = 'PO No.:'
ws['D13'].font = header_font
ws['D13'].alignment = right

ws.merge_cells('F13:G13')
ws['F13'] = ''  # PO No placeholder
ws['F13'].font = normal_font

# === Table Header ===
row = 15
headers = ['No.', 'Code', 'ITEM NAME / DESCRIPTION\n货名/摘要', 'QTY\n数量', 'UNIT\n单位', 'U_PRICE\n单价', 'AMOUNT($)\n金额']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    cell.fill = header_fill

ws.row_dimensions[row].height = 30

# === Data Rows (20 empty template rows) ===
for r in range(16, 36):
    ws.row_dimensions[r].height = 20
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c)
        cell.font = normal_font
        cell.border = thin_border
        if c == 1:  # No.
            cell.alignment = center
        elif c in (4, 5):  # QTY, UNIT
            cell.alignment = center
        elif c in (6, 7):  # Price, Amount
            cell.alignment = right
            cell.number_format = '#,##0.00'
        else:
            cell.alignment = left
    # Amount formula: QTY * U_PRICE
    ws.cell(row=r, column=7).value = f'=IF(D{r}<>"",D{r}*F{r},"")'

# === Summary Section ===
summary_row = 37
ws.merge_cells(f'A{summary_row}:E{summary_row}')
ws[f'F{summary_row}'] = '金额 SUB-TOTAL:'
ws[f'F{summary_row}'].font = header_font
ws[f'F{summary_row}'].alignment = right
ws[f'G{summary_row}'] = f'=SUM(G16:G35)'
ws[f'G{summary_row}'].font = header_font
ws[f'G{summary_row}'].alignment = right
ws[f'G{summary_row}'].number_format = '$#,##0.00'
ws[f'G{summary_row}'].border = thin_border

summary_row += 1
ws.merge_cells(f'A{summary_row}:E{summary_row}')
ws[f'F{summary_row}'] = '另加消费税9% Add GST 9%:'
ws[f'F{summary_row}'].font = header_font
ws[f'F{summary_row}'].alignment = right
ws[f'G{summary_row}'] = f'=G37*0.09'
ws[f'G{summary_row}'].font = header_font
ws[f'G{summary_row}'].alignment = right
ws[f'G{summary_row}'].number_format = '$#,##0.00'
ws[f'G{summary_row}'].border = thin_border

summary_row += 1
ws.merge_cells(f'A{summary_row}:E{summary_row}')
ws[f'F{summary_row}'] = '总金额 TOTAL:'
ws[f'F{summary_row}'].font = Font(name='Arial', size=12, bold=True)
ws[f'F{summary_row}'].alignment = right
ws[f'G{summary_row}'] = f'=G37+G38'
ws[f'G{summary_row}'].font = Font(name='Arial', size=12, bold=True)
ws[f'G{summary_row}'].alignment = right
ws[f'G{summary_row}'].number_format = '$#,##0.00'
ws[f'G{summary_row}'].border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='double')
)

# === Footer Notes ===
note_row = 41
ws.merge_cells(f'A{note_row}:G{note_row}')
ws[f'A{note_row}'] = 'Note: Any Shortage in delivery, missing of goods, please notify the company between 9am to 6pm otherwise we would assume they are received according to order.'
ws[f'A{note_row}'].font = Font(name='Arial', size=8)
ws[f'A{note_row}'].alignment = wrap

note_row += 1
ws.merge_cells(f'A{note_row}:G{note_row}')
ws[f'A{note_row}'] = 'All cheque should be crossed and made payable to DOLFORD FOOD MANUFACTURING PTE LTD'
ws[f'A{note_row}'].font = Font(name='Arial', size=8)

# Signature line
note_row += 2
ws.merge_cells(f'A{note_row}:C{note_row}')
ws[f'A{note_row}'] = "Customer's Signature & Company Stamp"
ws[f'A{note_row}'].font = small_font
ws[f'A{note_row}'].border = Border(top=Side(style='thin'))

# Print setup
ws.print_area = 'A1:G44'
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1

output_path = '/home/user/JoeyZzz/tax_invoice_template.xlsx'
wb.save(output_path)
print(f'Template saved to {output_path}')
