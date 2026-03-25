import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tax Invoice DFM8101729"

# ---- Styles ----
bold = Font(bold=True)
bold_large = Font(bold=True, size=14)
bold_medium = Font(bold=True, size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# ---- Column widths ----
col_widths = {"A": 5, "B": 12, "C": 50, "D": 8, "E": 8, "F": 10, "G": 12}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ---- Company Header ----
ws.merge_cells("A1:G1")
ws["A1"] = "豆福食品制造有限公司  DOLFORD FOOD MANUFACTURING PTE LTD"
ws["A1"].font = bold_large
ws["A1"].alignment = center

ws.merge_cells("A2:G2")
ws["A2"] = "8A Admiralty Street #02-17/18 FoodXchange@Admiralty (S)757437"
ws["A2"].alignment = center

ws.merge_cells("A3:G3")
ws["A3"] = "TEL:(65)65706728  HP:(65)91784466  FAX:(65)65706729  Email:newdolford@yahoo.com"
ws["A3"].alignment = center

ws.merge_cells("A4:G4")
ws["A4"] = "Company UEN No.:201010861E  OCBC:509733820001  GST Reg No:201010861E"
ws["A4"].alignment = center

# ---- Invoice Info ----
row = 6
ws.merge_cells(f"A{row}:D{row}")
ws[f"A{row}"] = "TAX INVOICE"
ws[f"A{row}"].font = bold_large
ws.merge_cells(f"E{row}:G{row}")
ws[f"E{row}"] = "INVOICE NO.: DFM8101729"
ws[f"E{row}"].font = bold_medium

row = 7
ws.merge_cells(f"E{row}:G{row}")
ws[f"E{row}"] = "DATE: 01/02/2026"
ws[f"E{row}"].font = bold_medium

# ---- Bill To / Deliver To ----
row = 9
ws.merge_cells(f"A{row}:C{row}")
ws[f"A{row}"] = "BILL TO:"
ws[f"A{row}"].font = bold
ws.merge_cells(f"E{row}:G{row}")
ws[f"E{row}"] = "DELIVER TO:"
ws[f"E{row}"].font = bold

bill_to = [
    "ZHANG LIANG MALA TANG",
    "张亮麻辣烫 (远东广场) (DELIVERY AFTER 10:00AM)",
    "FAR EAST PLAZA 远东广场",
    "14 SCOTTS RD #01-07A SINGAPORE 228213",
    "电话: 65310331"
]
for i, line in enumerate(bill_to):
    r = row + 1 + i
    ws.merge_cells(f"A{r}:C{r}")
    ws[f"A{r}"] = line
    ws.merge_cells(f"E{r}:G{r}")
    ws[f"E{r}"] = line  # Same deliver-to address

# ---- Customer Info ----
row = 16
ws[f"A{row}"] = "Customer ID: ZLML07"
ws[f"A{row}"].font = bold
ws[f"C{row}"] = "Term: 30 DAYS"
ws[f"E{row}"] = "Driver: 陈桥桥"
ws[f"G{row}"] = "P: 1 of 1"

# ---- Table Header ----
row = 18
headers = ["#", "Item Code", "ITEM NAME / DESCRIPTION 货名/摘要", "QTY 数量", "UNIT 单位", "U_PRICE 单价", "AMOUNT($) 金额"]
for i, h in enumerate(headers):
    cell = ws.cell(row=row, column=i+1, value=h)
    cell.font = bold
    cell.alignment = center
    cell.border = thin_border
    cell.fill = header_fill

# ---- Invoice Lines ----
items = [
    [1,  "BCS02",   "豆福豆皮（散装1KG）DOLFORD BEANCURD SKIN",           2.00,  "KGS", 3.60,  7.20],
    [2,  "CST01",   "芝士海鲜豆腐 CHEESE SEAFOOD TOFU",                   6.00,  "PKT", 3.50,  21.00],
    [3,  "ET01P",   "蛋豆腐 EGG TOFU 130g",                               6.00,  "PKT", 0.25,  1.50],
    [4,  "EZ01C",   "EZ快熟面 E-ZEE INSTANT NOODLES",                     1.00,  "CTN", 18.00, 18.00],
    [5,  "FD01P",   "炸豆枝片 EB FRIED BEAN CURD SHEET",                  2.00,  "PKT", 9.30,  18.60],
    [6,  "LB02P",   "FG 龙虾球 LOBSTER BALL (FG)",                        3.00,  "PKT", 3.50,  10.50],
    [7,  "NTU01",   "豆福手工卤水大豆腐 DOLFORD NORTH TOFU",               4.00,  "PCS", 1.20,  4.80],
    [8,  "SAT01",   "豆福五香豆干（散装）1KG DOLFORD SPICED",               1.00,  "KGS", 3.60,  3.60],
    [9,  "SDC01",   "烟熏鸭胸原味 Smoked Duck Original 5PKT/KG",           1.27,  "KGS", 9.50,  12.07],
    [10, "VTC02",   "豆福素鸡（散装）DOLFORD VEGETARIAN",                   1.00,  "KGS", 3.60,  3.60],
    [11, "WFB01",   "白鱼丸 (FG) 1KG*10PKT/CTN WHITE FISH",               1.00,  "KGS", 5.00,  5.00],
    [12, "ZSB0080", "金包银 MUSHROOM FISHN SOY 300G*30PKT/CTN",            2.00,  "PKT", 2.50,  5.00],
    [13, "NTU01",   "Exchange 交换: 豆福手工卤水大豆腐 DOLFORD",            4.00,  "PCS", None,  None],
    [14, "DB01P",   "豆卜 DARK BROWN BEAN CURD 10PCS X 200G",             15.00, "PKT", 1.40,  21.00],
]

for i, item in enumerate(items):
    r = row + 1 + i
    for j, val in enumerate(item):
        cell = ws.cell(row=r, column=j+1, value=val)
        cell.border = thin_border
        if j in (0, 3, 4, 5, 6):
            cell.alignment = center
        else:
            cell.alignment = left_wrap
        # Format numbers
        if j in (3, 5, 6) and val is not None:
            cell.number_format = '#,##0.00'

# ---- Totals ----
total_row = row + 1 + len(items) + 1
ws.merge_cells(f"E{total_row}:F{total_row}")
ws[f"E{total_row}"] = "金额 SUB-TOTAL:"
ws[f"E{total_row}"].font = bold
ws[f"E{total_row}"].alignment = Alignment(horizontal="right")
ws[f"G{total_row}"] = 131.87
ws[f"G{total_row}"].number_format = '$#,##0.00'
ws[f"G{total_row}"].font = bold
ws[f"G{total_row}"].alignment = center

total_row += 1
ws.merge_cells(f"E{total_row}:F{total_row}")
ws[f"E{total_row}"] = "另加消费税9% Add GST 9%:"
ws[f"E{total_row}"].font = bold
ws[f"E{total_row}"].alignment = Alignment(horizontal="right")
ws[f"G{total_row}"] = 11.87
ws[f"G{total_row}"].number_format = '$#,##0.00'
ws[f"G{total_row}"].font = bold
ws[f"G{total_row}"].alignment = center

total_row += 1
ws.merge_cells(f"E{total_row}:F{total_row}")
ws[f"E{total_row}"] = "总金额 TOTAL:"
ws[f"E{total_row}"].font = Font(bold=True, size=13)
ws[f"E{total_row}"].alignment = Alignment(horizontal="right")
ws[f"G{total_row}"] = 143.74
ws[f"G{total_row}"].number_format = '$#,##0.00'
ws[f"G{total_row}"].font = Font(bold=True, size=13)
ws[f"G{total_row}"].alignment = center

# ---- Notes ----
total_row += 2
ws.merge_cells(f"A{total_row}:G{total_row}")
ws[f"A{total_row}"] = "AMENDED"
ws[f"A{total_row}"].font = Font(bold=True, size=14, color="FF0000")
ws[f"A{total_row}"].alignment = center

total_row += 1
ws.merge_cells(f"A{total_row}:G{total_row}")
ws[f"A{total_row}"] = "XIAOZHENG"

total_row += 1
ws.merge_cells(f"A{total_row}:G{total_row}")
ws[f"A{total_row}"] = "01/02/2026 07:40:15 AM"

# Save
output_path = "/home/user/JoeyZzz/invoice_DFM8101729.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
